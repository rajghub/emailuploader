# --------- EXTRACT DATA FROM EXCEL FILE ------------------------
import openpyxl

class accessExcel:
    book = ""
    sheet = ""
    fieldDetails = {}
    cellCoordinates = {}

# Initialisation
    def __init__(self, exlPath, sheetName):
        self.exlPath = exlPath
        self.sheetName = sheetName

# Open Excel and select sheet
    def selectSheet(self):
        self.book = openpyxl.load_workbook(self.exlPath)
        self.sheet = self.book[self.sheetName]


# Access and collect required fields
    def accessFields(self, valuerow):

        self.headingrow=2
        self.valuerow = valuerow

        for i in range(1, self.sheet.max_column+1):
            hr=self.sheet.cell(self.headingrow, i)
            vr=self.sheet.cell(self.valuerow, i)
            self.fieldDetails[hr.value] = vr.value
            # if(hr.value == "Ref No"):
            #     self.fieldDetails["Refno_cell"]=vr.coordinate
            if(hr.value == "Tracking Id"):
                self.fieldDetails["trackingid_cell"]=vr.coordinate

        return self.fieldDetails

#   To modify the sheet
    def modifySheet(self, originText, textToReplace, replaceWith, cellNo):
        self.originText = originText
        self.originText = self.originText.replace(textToReplace, replaceWith)
        self.fieldDetails["Tracking Id"]=self.originText
        self.sheet[cellNo] = self.originText
        self.book.save(self.exlPath)

    def updateDocInfo(self,docInfoNumber):
        self.docInfoNumber=docInfoNumber
        # print(self.fieldDetails["Refno_cell"])
        self.sheet['C1'] = self.docInfoNumber
        self.book.save(self.exlPath)

    def updateCell(self, pardDetails):
        self.refCellValue=self.sheet[pardDetails["source_cell"]].value
        self.fieldDetails["Tracking Id"] = "{{$"+str(self.refCellValue)+"}}&amp;"+self.sheet[self.fieldDetails["trackingid_cell"]].value
        print(self.fieldDetails["Tracking Id"])
        self.sheet[pardDetails["Dest_cell"]]=self.fieldDetails["Tracking Id"]
        self.book.save(self.exlPath)