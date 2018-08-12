import openpyxl

wb = openpyxl.load_workbook("uploadData.xlsx")
ws = wb.active
counter =1


cols = ws.max_column
headingrow=2
valuerow =4
fieldDetails={}

for i in range(1, cols+1):
    
    hr=ws.cell(headingrow, i)
    # print(hr.coordinate)
    vr=ws.cell(valuerow, i)
    # print("--------------------------------")
    # print(vr.coordinate)
    fieldDetails[hr.value] = vr.value
    if(hr.value == "Ref No"):
        print(hr.coordinate)
        print(vr.coordinate)

# for rowOfCellObjects in ws['A1':'C3']:
#     for cellObj in rowOfCellObjects:
#         print(cellObj.coordinate, cellObj.value)
#     print('--- END OF ROW ---')


# for i in range(1, cols+1):
#     for cellObj in i:
#         print(cellObj.coordinate, cellObj.value)
    
    # hr=ws.cell(headingrow, i)
    # vr=ws.cell(valuerow, i)
    # fieldDetails[hr.value] = vr.value
# print(fieldDetails)
# for cell in cellRange:
    
#     cnumber='A'+str(counter)
#     print(cnumber)
#     print(ws[cnumber].value)
#     counter+=1
